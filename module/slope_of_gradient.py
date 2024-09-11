# NMFF-AFM: normal mode flexible fitting of proteins conformations to AFM images
#
# Copyright (c) 2024 TamaLab
# Authors: Xuan Wu, Nagoya University
#          Osamu Miyashita, RIKEN
#          Florence Tama, Nagoya University/RIKEN
#
# References:
# Modeling Conformational Transitions of Biomolecules from Atomic Force
# Microscopy Images using Normal Mode Analysis
# Xuan Wu, Osamu Miyashita, and Florence Tama
# https://doi.org/10.1021/acs.jpcb.4c04189
#
#
import os
import openpyxl
import xlsxwriter
import numpy as np
import pandas as pd
from pathlib import Path
from scipy.stats import pearsonr
from sklearn.linear_model import LinearRegression


def create_step_workbook(path_of_file, amplitudes, start_mode, end_mode):
    """
    This function will create a log for the current step
    :param path_of_file: The path to the Excel file
    :param amplitudes: The amplitude used in the flexible fit-in
    :param start_mode: From which mode will be used in the flexible fit-in
    :param end_mode: Till which mode will be used in the flexible fit-in
    :return: None
    """
    workbook = xlsxwriter.Workbook(path_of_file)
    worksheet = workbook.add_worksheet()

    # Write the 'numbers' of the amplitude for Pearson table in the sheet
    for i, num in enumerate(amplitudes):
        worksheet.write(0, i + 1, num)

    values = []
    for num in range(start_mode, end_mode + 1, 1):
        values.append(f"Mode {str(num)}")

    # Write 'Modes' for Pearson plot table
    for i, value in enumerate(values):
        worksheet.write(i + 1, 0, value)

    workbook.close()

    print("Blank table created.")


def write_to_xlsx(io, df_cc):
    """
    Write the plots calculated by Pearson into the specific cells
    :param io: The path to the Excel file
    :param df_cc: The list contains CC at each amplitude
    :return:
    """
    wb = openpyxl.load_workbook(io)
    sheet = wb.active
    df_cc_pivot = df_cc.pivot(columns="dq", index="mode")
    (nr, nc) = df_cc_pivot.shape
    for r in range(nr):
        for c in range(nc):
            cell = sheet.cell(row=r + 2, column=c + 2)
            cell.value = df_cc_pivot.iloc[r, c]
    wb.save(io)


def slope_write_to_xlsx(io, df_slopes, n_amplitudes):
    """
    Write the slopes calculated from Pearson plots into the specific cells
    :param io: The path to the log
    :param df_slopes: The list contains slope and intercept of one mode
    :param n_amplitudes: The number of how many frequencies used in the flexible fit-in
    :return: None
    """
    wb = openpyxl.load_workbook(io)
    sheet = wb.active
    nrows = len(df_slopes.index)
    for r in range(nrows):
        cell = sheet.cell(row=r + 2, column=n_amplitudes + 2)
        cell.value = df_slopes["slope"].iloc[r]
        cell = sheet.cell(row=r + 2, column=n_amplitudes + 3)
        cell.value = df_slopes["intercept"].iloc[r]
    wb.save(io)


def slope_calculation(df):
    """
    This function will calculate the slope with the input plots with linear function
    :param df: A list contains the value at each amplitude
    :return: A list contains the gradient and the intercept
    """
    x = df[["dq"]]  # Input features (independent variable)
    y = df["cc"]  # Target variable (dependent variable)

    # Create and fit the linear regression model
    model = LinearRegression()
    model.fit(x, y)

    slope = float(model.coef_[0])
    intercept = float(model.intercept_)

    return [slope, intercept]


def find_largest_slopes(df):
    """
    This function will find the largest and second-largest slope among all the slopes then return them
    :param df: The input dictionary contains the gradients
    :return: The modes with first and second-largest gradient
    """
    largest_slope = float(0)
    second_largest_slope = float(0)
    largest_name = None
    second_largest_name = None

    df['abs_slope'] = df['slope'].apply(abs)
    df = df.sort_values(by=['abs_slope'], ascending=False)

    largest_mode = df['mode'].iloc[0]
    largest_slope = df['slope'].iloc[0]
    second_largest_mode = df['mode'].iloc[1]
    second_largest_slope = df['slope'].iloc[1]

    return largest_mode, largest_slope, second_largest_mode, second_largest_slope


def read_tsv(filename):
    """
    This function will read the TSV file and convert them into matrix
    :param filename: The file should be converted
    :return: Matrix representation of the image
    """
    mat = np.loadtxt(filename)
    return mat


def calculation_correlation_pearson_tsv_normalized(target, initial):
    """
    This function will calculate the CC between the two figures
    :param target: The path to the target figure
    :param initial: The path to the initial figure
    :return: Float cc
    """
    pmat = read_tsv(target)
    imat = read_tsv(initial)
    shifted_pmat = pmat - pmat.min()
    shifted_imat = imat - imat.min()
    normalized_pmat = shifted_pmat / pmat.max()
    normalized_imat = shifted_imat / imat.max()
    normalized_pmat_flatten = normalized_pmat.flatten()
    normalized_imat_flatten = normalized_imat.flatten()
    cc = pearsonr(normalized_pmat_flatten, normalized_imat_flatten)[0]
    return cc


def cc_calculation_pearson_tsv(
        directory, target_figure_path, initial_conformation_name, fig_mode
):
    """
    This function will calculate the CC between the reference figure and all the other figures in TSV file
    in the same folder using PearsonR, usually named /1st
    :param directory: The path to the directory containing the figures
    :param target_figure_path: The path to the target figure file used as a reference for CC calculation
    :param initial_conformation_name: The name of the initial conformation without suffix
    :param fig_mode: The type of the figures used in the flexible fit-in
    :return: pd.DataFrame: A DataFrame containing the mode, dq (amplitude), and CC values
    """
    # Create an empty DataFrame with columns mode, dq, and cc
    df = pd.DataFrame(columns=["mode", "dq", "cc"])

    # Iterate over all files in the specified directory and calculate CC
    for filename in os.listdir(directory):
        if filename.endswith(f".{fig_mode}"):
            if "#" in filename:
                if initial_conformation_name not in filename:
                    initial_file_path = os.path.join(directory, filename)
                    cc = float(
                        calculation_correlation_pearson_tsv_normalized(
                            target_figure_path, initial_file_path
                        )
                    )
                    mode_number = filename.split("#")[0]
                    amplitude = filename.split("#")[1].split(".t")[0]
                    df.loc[len(df.index)] = [mode_number, amplitude, cc]

    # Try to convert mode and dq columns to integers and cc column to float
    try:
        df["mode"] = df["mode"].astype("int")
        df["dq"] = df["dq"].astype("int")
        df["cc"] = df["cc"].astype("float")
    except ValueError as e:
        # If astype conversion fails, print the DataFrame
        print(f"tsv file import failed.")
        print(df)
        print(f"Error message: {e}")

    df.to_csv(Path(directory) / "cc_table.csv")

    return df


def process_data_tsv(
        step_workbook_path,
        reference_fig,
        original_conformation_name,
        start_mode,
        end_mode,
        fig_usage,
        figure_directory
):
    """
    This function will calculate the CC and the gradient and write them into the log for current step
    :param step_workbook_path: The path to the log of current step
    :param reference_fig: The path to the target figure
    :param original_conformation_name: The name of the initial conformation without suffix
    :param start_mode: The number from which the frequency will be used in the flexible fit-in
    :param end_mode: The number till which the frequency will be used in the flexible fit-in
    :param fig_usage: Which format is using in the flexible fit-in
    :param figure_directory: The path to the folder contains all the simulated figures
    :return df_slopes, df_pcc_sorted: Lists of CCs and slopes
    """
    if os.path.exists(step_workbook_path):
        raise FileExistsError(f"{step_workbook_path} already exists")

    df_pcc = cc_calculation_pearson_tsv(
        figure_directory, reference_fig, original_conformation_name, fig_usage
    )

    df_slopes = pd.DataFrame(columns=["mode", "slope", "intercept"])
    for i in range(start_mode, end_mode + 1):
        df_pcc_one_mode = df_pcc[df_pcc["mode"] == i].sort_values(by=["dq"])
        slope = slope_calculation(df_pcc_one_mode)
        df_slopes.loc[len(df_slopes)] = [i, slope[0], slope[1]]

    amplitudes = df_pcc['dq'].unique()
    amplitudes.sort()
    create_step_workbook(step_workbook_path, amplitudes, start_mode, end_mode)

    write_to_xlsx(step_workbook_path, df_pcc)
    slope_write_to_xlsx(step_workbook_path, df_slopes, len(amplitudes))

    df_slopes["mode"] = df_slopes["mode"].astype("int")

    df_pcc_sorted = df_pcc.sort_values(by=['cc'], ascending=False)

    return df_slopes, df_pcc_sorted
