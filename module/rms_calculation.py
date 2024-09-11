# NMFF-AFM: normal mode flexible fitting of proteins conformations to AFM images
#
# Copyright (c) 2024 TamaLab
# Authors: Xuan Wu, Nagoya University
#          Osamu Miyashita, RIKEN
#          Florence Tama, Nagoya University/RIKEN
#
# References:
# Modeling Conformational Transitions of Biomolecules from Atomic Force
# Microscopy Images using Normal Mode Analysis
# Xuan Wu, Osamu Miyashita, and Florence Tama
# https://doi.org/10.1021/acs.jpcb.4c04189
#
#
import os
import re
import shutil
import openpyxl
import subprocess
from pathlib import Path

profit_path = os.environ.get('PRO_FIT_PATH')


def rmsd_calculation(folder_path, reference_pdb):
    """
    This function will calculate the RMSD value between all the PDB files in the folder and the reference PDB file
    :param folder_path: The **PATH** of the folder which contains all the PDB files used as reference and mobile
    :param reference_pdb: The **NAME** of the reference PDB file
    :return: None
    """
    # Prepare an unfinished Pro-Fit script with only REFERENCE and MOBILE files, every 50 files in one script
    files = os.listdir(folder_path)
    for i in range(0, len(files), 50):
        chunk = files[i:i+50]
        with open(Path(folder_path) / f"script_{i//50 + 1}.txt", 'w') as f:
            f.write(f'reference {reference_pdb}.pdb' + '\n')
            for file_name in chunk:
                if file_name.endswith('.pdb'):
                    if not file_name.split('.')[0] == reference_pdb:
                        f.write(f'mobile {file_name}' + '\n')
                        f.write('fit' + '\n')
            f.write('quit\n')

    print("Scripts generated.")

    # Calculate the RMSD using Pro-Fit
    for script_name in os.listdir(folder_path):
        if script_name.endswith('.txt'):
            if 'script' in script_name:
                # result_file_path = os.path.join(folder_path, "result-" + script_name.split('_')[1])
                result_file_path = f"result-{script_name.split('_')[1]}"
                cmd = f"cd {folder_path} && {profit_path} -h -f {script_name} &> {result_file_path}"
                subprocess.run(cmd, shell=True, check=True)

    print("RMSD calculation finished and results written to text file.")


def write_rmsd_to_cell_by_keyword(io_path, keyword, result_rmsd, column_num):
    """
    This function will write the given value into the specific cell of the Excel file by searching the content of the 1st cell of the row
    :param column_num: The number of the column to put the value of RMSD
    :param io_path: The PATH of the Excel file
    :param keyword: The one should be searched in the 1st cell of the row
    :param result_rmsd: The value of RMSD should be written into the cell at the end of the row(5th cell)
    :return: None
    """
    wb = openpyxl.load_workbook(io_path)
    ws = wb.active

    for row in ws.iter_rows(1):
        for cell in row:
            if cell.value == keyword:
                ws.cell(row=cell.row, column=column_num, value=result_rmsd)

    # Write the initial RMSD 0.00 into logfile
    ws.cell(row=2, column=6, value=0.000)
    wb.save(io_path)


def write_rmsd_to_dictionary(rms_result_txt):
    """
    This function will use regular equation to put all the name of the PDB files and the related RMSD into a dictionary
    :param rms_result_txt: The path to the result file
    :return: A dictionary contains the formatted results
    """
    with open(rms_result_txt, "r") as f:
        file_lines = f.readlines()

    pattern_pdb = r"Reading mobile structure \((.*)\)"
    pdb_files = []

    pattern_rms = r"RMS: \d+\.\d+"
    rms_values = []

    for line in file_lines:
        matches = re.findall(pattern_pdb, line)
        for match in matches:
            match = match.split(".pdb")[0]
            pdb_files.append(match)

    for line in file_lines:
        matches = re.findall(pattern_rms, line)
        for match in matches:
            match = match.split(": ")[1]
            rms_values.append(float(match))

    dictionary = dict(zip(pdb_files, rms_values))
    return dictionary


def process_rmsd_dictionary(output_excel, pdb_folder, num_of_column):
    """
    This function will read the result text files and write them into the cc_changing Excel file
    :param num_of_column:
    :param output_excel: The path to the cc_changing.xlsx
    :param pdb_folder: Path to All_conformation
    :return: None
    """
    rmsd_dictionary = {}

    for files in os.listdir(pdb_folder):
        if "result" in files:
            rmsd_dictionary = write_rmsd_to_dictionary(Path(pdb_folder) / files)

        for key_word in rmsd_dictionary:
            value = rmsd_dictionary[key_word]
            write_rmsd_to_cell_by_keyword(output_excel, key_word, value, num_of_column)

    print("RMSD calculation finished, all data written to log.")


def rmsd_to_initial_and_reference(upper_folder_path, initial_conformation_name, reference_pdb_name, overall_log_file_path, whether_calculate_rmsd_reference):
    """
    This function will calculate RMSD to Initial/Reference basing on setting
    :param upper_folder_path: The path t the folder contains all the subfolders
    :param initial_conformation_name: The name of the initial conformation
    :param reference_pdb_name: The name of the reference conformation
    :param overall_log_file_path: The path to the overall log
    :param whether_calculate_rmsd_reference: "yes" or "no" for deciding whether to calculate the RMSD-Reference
    :return: None
    """

    if whether_calculate_rmsd_reference == "yes":
        # Copy reference PDB into /All_conformation
        shutil.copy(Path(upper_folder_path) / f"{reference_pdb_name}.pdb",
                    Path(upper_folder_path) / "All_conformation" / f"{reference_pdb_name}.pdb")

        # Calculate RMSD to initial conformation and write it into log file
        rmsd_calculation(Path(upper_folder_path) / "All_conformation", initial_conformation_name)
        process_rmsd_dictionary(overall_log_file_path, Path(upper_folder_path) / "All_conformation", 6)

        # Calculate RMSD to reference conformation and write it into log file
        rmsd_calculation(Path(upper_folder_path) / "All_conformation", reference_pdb_name)
        process_rmsd_dictionary(overall_log_file_path, Path(upper_folder_path) / "All_conformation", 7)

        print(
            "RMSD to Initial conformation and Reference conformation have been finished and results written to logfile."
        )
    else:
        # Calculate RMSD to initial conformation and write it into log file
        rmsd_calculation(Path(upper_folder_path) / "All_conformation", initial_conformation_name)
        process_rmsd_dictionary(overall_log_file_path, Path(upper_folder_path) / "All_conformation", 6)

        print(
            "RMSD to Initial conformation has been finished and results written to logfile, RMSD to Reference conformation has been skipped."
        )
