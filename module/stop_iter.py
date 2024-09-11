# NMFF-AFM: normal mode flexible fitting of proteins conformations to AFM images
#
# Copyright (c) 2024 TamaLab
# Authors: Xuan Wu, Nagoya University
#          Osamu Miyashita, RIKEN
#          Florence Tama, Nagoya University/RIKEN
#
# References:
# Modeling Conformational Transitions of Biomolecules from Atomic Force
# Microscopy Images using Normal Mode Analysis
# Xuan Wu, Osamu Miyashita, and Florence Tama
# https://doi.org/10.1021/acs.jpcb.4c04189
#
#
import os
import openpyxl
from pathlib import Path

from module.slope_of_gradient import calculation_correlation_pearson_tsv_normalized


def create_log(excel_file_path, whether_contain_rmsd_reference):
    """
    This function will create an empty Excel file with header as the log for the whole flexible fit-in
    :param excel_file_path:
    :param whether_contain_rmsd_reference: Whether to calculate the RMSD-Reference
    :return: None
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    if whether_contain_rmsd_reference == "yes":
        # Add headers if the Excel file is newly created
        ws.append([
            "Deformed fig name", "CC of this iter", "Largest mode", "Amplitude", "Last 5 iter average CC",
            "RMSD to Initial",
            "RMSD to Reference"
        ])
    else:
        # Add headers if the Excel file is newly created
        ws.append([
            "Deformed fig name", "CC of this iter", "Largest mode", "Amplitude", "Last 5 iter average CC",
            "RMSD to Initial"
        ])

    # Save the changes to the Excel file
    wb.save(excel_file_path)
    print("Log has been created.")


def write_log(io, record):
    """
    This function will write the list contains the data for this step into the log
    :param io: The path to the overall log
    :param record: The list contains the data
    :return: None
    """
    wb = openpyxl.load_workbook(io)
    ws = wb.active
    ws.append(record)
    wb.save(io)
    print("Data of this step written to the log.")


def read_pre_iter_cc_and_avrg(excel_file_path, required_num):
    """
    This function will read the CC from the Excel file and return an average value as required
    :param excel_file_path: The path to the overall log
    :param required_num: The number of howe many lines should be read from the N-5 line from current line
    :return: The sum of the required lines
    """
    # Load the Excel file
    workbook = openpyxl.load_workbook(excel_file_path)

    # Select the desired worksheet
    worksheet = workbook['Sheet']

    # Find the last row in the worksheet
    last_row = worksheet.max_row

    total_value = 0

    # Read the values in the last row into an array
    for i in range(0, required_num, 1):
        last_row_values = [worksheet.cell(row=last_row - i, column=column).value for column in
                           range(1, worksheet.max_column + 1)]

        # Select the CC value of the last iteration
        last_value = last_row_values[1]
        total_value = total_value + last_value

    # Return the array representing the last row
    return total_value


def cc_judgement_avrg(work_folder, deformed_fig_name, target_fig_name, mode_max_gradient, deform_amplitude,
                      excel_file_path, suffix_of_fig):
    """
    This function will compare the CC of this iteration to the last one and judge the iteration should be stopped
    or not, then write the CC and the other parameters of the new iteration into the same Excel file
    :param work_folder: The path to the folder contains all the files
    :param deformed_fig_name: The name of the initial figure of this step
    :param target_fig_name: The name of the target figure
    :param mode_max_gradient: The number of the mode with the max gradient
    :param deform_amplitude: The value of the amplitude used in this step to deform the PDB for next step
    :param excel_file_path: The path to the overall log
    :param suffix_of_fig: The suffix of the figure used in the flexible fit-in
    :return: The bool of whether to stop the iteration
    """
    target_fig_path = Path(os.path.dirname(excel_file_path)) / f"{target_fig_name}.{suffix_of_fig}"
    deformed_fig_path = Path(work_folder) / f"{deformed_fig_name}.{suffix_of_fig}"

    try:
        cc_last_5_iter_avrg = read_pre_iter_cc_and_avrg(excel_file_path, 5) / 5
    except Exception as e:
        cc_last_5_iter_avrg = 0

    cc_iter = calculation_correlation_pearson_tsv_normalized(target_fig_path, deformed_fig_path)

    try:
        cc_iter_this_iter_avrg = (read_pre_iter_cc_and_avrg(excel_file_path, 4) + cc_iter) / 5
    except Exception as e:
        cc_iter_this_iter_avrg = cc_iter

    record_list = [deformed_fig_name, cc_iter, mode_max_gradient, deform_amplitude, cc_iter_this_iter_avrg]
    write_log(excel_file_path, record_list)

    # Return the result of the judgement
    if cc_iter_this_iter_avrg > cc_last_5_iter_avrg:
        return True
    else:
        return False


def read_pre_iter_cc(excel_file_path):
    """
    This function will read the CC from the Excel file and return it
    :param excel_file_path: The path to the overall log
    :return: The CC value of last record
    """

    # Load the Excel file
    workbook = openpyxl.load_workbook(excel_file_path)

    # Select the desired worksheet
    worksheet = workbook['Sheet']

    # Find the last row in the worksheet
    last_row = worksheet.max_row

    # Read the values in the last row into an array
    last_row_values = [worksheet.cell(row=last_row, column=column).value for column in
                       range(1, worksheet.max_column + 1)]

    # Select the CC value of the last iteration
    last_value = last_row_values[1]

    # Return the array representing the last row
    return last_value


def cc_judgement_single(work_folder, deformed_fig_name, target_fig_name, mode_max_gradient, deform_amplitude,
                        excel_file_path, suffix_of_fig):
    """
    This function will compare the CC of this iteration to the last one and judge the iteration should be stopped
    or not, then write the CC and the other parameters of the new iteration into the same Excel file
    :param work_folder: The path to the folder contains all the files
    :param deformed_fig_name: The name of the initial figure of this step
    :param target_fig_name: The name of the target figure
    :param mode_max_gradient: The number of the mode with the max gradient
    :param deform_amplitude: The value of the amplitude used in this step to deform the PDB for next step
    :param excel_file_path: The path to the overall log
    :param suffix_of_fig: The suffix of the figure used in the flexible fit-in
    :return: The bool of whether to stop the iteration
    :param suffix_of_fig: The suffix of the figure used in the flexible fit-in
    :return: The bool of whether to stop the iteration
    """

    reference_fig_path = f"{work_folder}{target_fig_name}.{suffix_of_fig}"
    deformed_fig_path = f"{work_folder}{deformed_fig_name}.{suffix_of_fig}"

    try:
        cc_last_iter = read_pre_iter_cc(excel_file_path)
    except FileNotFoundError:
        cc_last_iter = 0

    cc_iter = calculation_correlation_pearson_tsv_normalized(reference_fig_path, deformed_fig_path)

    record_list = [deformed_fig_name, cc_iter, mode_max_gradient, deform_amplitude]
    write_log(excel_file_path, record_list)

    # Return the result of the judgement
    if cc_iter > cc_last_iter:
        return True
    else:
        return False
